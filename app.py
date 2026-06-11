"""
UNIAGRARIA - SISTEMA DE RECOLECCIÓN FACATATIVÁ 2026
Julian Camilo Quintero Martinez
VERSIÓN COMPLETA CON MÓDULO WHATSAPP INTEGRADO Y AUTENTICACIÓN MICROSOFT
"""

import os
import json
import csv
import zipfile
import secrets
import re
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

# Flask y extensiones core
from flask import (
    Flask, render_template, request, redirect, url_for,
    session, jsonify, flash, send_file, make_response
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, current_user, login_required
from flask_migrate import Migrate
from flask_cors import CORS
from flask_compress import Compress
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_caching import Cache
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix

# Base de datos
from sqlalchemy import func

# Cloudinary
import cloudinary
import cloudinary.uploader
import cloudinary.api

# Procesamiento de datos
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# PDF y Reportes
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Utilidades
from email_validator import validate_email, EmailNotValidError

# MSAL para Azure AD
from msal import ConfidentialClientApplication

# Configuración de variables de entorno
from dotenv import load_dotenv
load_dotenv()

# ============================================================================
# CONFIGURACIÓN MICROSOFT AZURE AD
# ============================================================================

AZURE_CLIENT_ID = os.environ.get('AZURE_CLIENT_ID', '')
AZURE_CLIENT_SECRET = os.environ.get('AZURE_CLIENT_SECRET', '')
AZURE_TENANT_ID = os.environ.get('AZURE_TENANT_ID', 'common')
AZURE_AUTHORITY = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"
AZURE_SCOPE = ["User.Read", "email", "openid", "profile"]

# Inicializar MSAL solo si hay configuración
msal_app = None
if AZURE_CLIENT_ID and AZURE_CLIENT_SECRET:
    try:
        msal_app = ConfidentialClientApplication(
            client_id=AZURE_CLIENT_ID,
            client_credential=AZURE_CLIENT_SECRET,
            authority=AZURE_AUTHORITY
        )
        print("✅ Microsoft Azure AD configurado")
    except Exception as e:
        print(f"⚠️ Error configurando Microsoft Azure AD: {e}")
else:
    print("⚠️ Microsoft Azure AD no configurado - solo login por email")

# ============================================================================
# CONFIGURACIÓN INICIAL DE LA APLICACIÓN
# ============================================================================

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# Configuración de sesión
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['SESSION_COOKIE_NAME'] = 'session'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_REFRESH_EACH_REQUEST'] = True

# ==============================================
# CONFIGURACIÓN NEONTECH POSTGRESQL
# ==============================================
database_url = os.environ.get('DATABASE_URL', '')

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv(
    "DATABASE_URL",
    "sqlite:///test.db"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_size': 10,
    'pool_recycle': 300,
    'pool_pre_ping': True,
    'pool_use_lifo': True,
    'max_overflow': 20,
    'connect_args': {
        'sslmode': 'require',
        'connect_timeout': 10,
        'keepalives': 1,
        'keepalives_idle': 30,
        'keepalives_interval': 10,
        'keepalives_count': 5
    }
}

# Configuración de archivos
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024  # 500MB
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf', 'xlsx', 'xls', 'csv'}
app.config['ALLOWED_IMAGES'] = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# Crear carpetas necesarias
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('static/plantillas', exist_ok=True)

# ==============================================
# CONFIGURACIÓN CLOUDINARY
# ==============================================
CLOUDINARY_CLOUD_NAME = os.environ.get('CLOUDINARY_CLOUD_NAME', '')
CLOUDINARY_API_KEY    = os.environ.get('CLOUDINARY_API_KEY', '')
CLOUDINARY_API_SECRET = os.environ.get('CLOUDINARY_API_SECRET', '')

print("=" * 50)
print("🔵 CONFIGURACIÓN CLOUDINARY")
print(f"🔵 Cloud Name: {CLOUDINARY_CLOUD_NAME}")
print(f"🔵 API Key: {CLOUDINARY_API_KEY[:5]}..." if CLOUDINARY_API_KEY else "🔵 API Key: NO CONFIGURADA")
print("=" * 50)

CLOUDINARY_CONFIGURED = False
try:
    if CLOUDINARY_CLOUD_NAME and CLOUDINARY_API_KEY and CLOUDINARY_API_SECRET:
        cloudinary.config(
            cloud_name=CLOUDINARY_CLOUD_NAME,
            api_key=CLOUDINARY_API_KEY,
            api_secret=CLOUDINARY_API_SECRET,
            secure=True
        )
        test_result = cloudinary.api.ping()
        print(f"✅ CLOUDINARY CONECTADO: {test_result}")
        CLOUDINARY_CONFIGURED = True
    else:
        print("⚠️ Cloudinary no configurado. Las imágenes se guardarán localmente.")
except Exception:
    CLOUDINARY_CONFIGURED = False
    print("❌ Error Cloudinary")

# Inicializar extensiones
db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder'
login_manager.session_protection = 'strong'
CORS(app, supports_credentials=True)
Compress(app)
cache = Cache(app, config={'CACHE_TYPE': 'SimpleCache', 'CACHE_DEFAULT_TIMEOUT': 300})

limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    strategy="fixed-window"
)

# ============================================================================
# FUNCIÓN PARA VALIDAR SI UN CONTACTO ES VÁLIDO (NO ES SISTEMA)
# ============================================================================

def es_contacto_valido(contacto):
    """Filtra contactos que NO son del sistema (status, grupos, IDs internos)"""
    if not contacto:
        return False
    # Excluir mensajes de sistema
    if contacto == 'status@broadcast':
        return False
    # Excluir IDs de grupos (terminan en @g.us)
    if contacto.endswith('@g.us'):
        return False
    # Excluir IDs internos de WhatsApp (contienen @lid, etc)
    if '@' in contacto and not contacto.endswith('@c.us'):
        return False
    # Debe ser un número de teléfono válido (solo dígitos, + y -)
    telefono_limpio = re.sub(r'[^0-9]', '', contacto)
    if len(telefono_limpio) < 8:
        return False
    return True

def limpiar_numero_contacto(contacto):
    """Limpia el número de contacto para mostrar (NO elimina @lid porque son contactos válidos)"""
    if not contacto:
        return None
    # Eliminar sufijo @c.us si existe
    if contacto.endswith('@c.us'):
        contacto = contacto[:-5]
    # Eliminar otros sufijos (pero NO @lid)
    contacto = contacto.replace('@s.whatsapp.net', '')
    contacto = contacto.replace('@g.us', '')
    # Si es status, retornar None
    if contacto == 'status@broadcast':
        return None
    # NO eliminar @lid - son contactos válidos de WhatsApp
    # Verificar que tenga al menos 3 caracteres (los @lid pueden ser cortos)
    if len(contacto) < 3:
        return None
    return contacto

# ============================================================================
# CONFIGURACIÓN WHATSAPP
# ============================================================================

# Mapeo número → clave interna
WA_ASESORAS = {
    '573212526461': 'asesora1',   # Karen Valencia Pinto Ladino
    '573107694751': 'asesora2',   # Ingrid
}

# Nombres para mostrar en la interfaz
WA_NOMBRES = {
    'asesora1': 'Karen Valencia Pinto Ladino',
    'asesora2': 'Ingrid Del Rosario Maldonado Gonzales',
}

# ============================================================================
# MODELOS DE BASE DE DATOS
# ============================================================================

class Usuario(UserMixin, db.Model):
    __tablename__ = 'usuarios'

    id             = db.Column(db.Integer, primary_key=True)
    email          = db.Column(db.String(255), unique=True, nullable=False, index=True)
    nombre         = db.Column(db.String(255), nullable=False)
    apellido       = db.Column(db.String(255))
    azure_id       = db.Column(db.String(255), unique=True)
    rol            = db.Column(db.String(50), default='usuario')
    activo         = db.Column(db.Boolean, default=True)
    avatar_url     = db.Column(db.String(500))
    ultimo_acceso  = db.Column(db.DateTime)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)

    registros        = db.relationship('RecoleccionDato', back_populates='usuario_registro', lazy=True)
    imagenes_subidas = db.relationship('FeriaImagen',     back_populates='usuario_subida',   lazy=True)
    importaciones    = db.relationship('ArchivoImportado', back_populates='usuario_importo', lazy=True)

    def is_admin(self):
        return self.rol == 'admin'

    def to_dict(self):
        return {
            'id':            self.id,
            'email':         self.email,
            'nombre':        self.nombre,
            'apellido':      self.apellido,
            'rol':           self.rol,
            'activo':        self.activo,
            'avatar':        self.avatar_url,
            'ultimo_acceso': self.ultimo_acceso.isoformat() if self.ultimo_acceso else None,
            'fecha_registro': self.fecha_registro.isoformat() if self.fecha_registro else None
        }


class Municipio(db.Model):
    __tablename__ = 'municipios'

    id            = db.Column(db.Integer, primary_key=True)
    nombre        = db.Column(db.String(100), unique=True, nullable=False, index=True)
    departamento  = db.Column(db.String(100), default='Cundinamarca')
    activo        = db.Column(db.Boolean, default=True)

    instituciones = db.relationship('Institucion',     back_populates='municipio', lazy=True)
    registros     = db.relationship('RecoleccionDato', back_populates='municipio', lazy=True)

    def to_dict(self):
        return {'id': self.id, 'nombre': self.nombre, 'departamento': self.departamento}


class Institucion(db.Model):
    __tablename__ = 'instituciones'

    id           = db.Column(db.Integer, primary_key=True)
    nombre       = db.Column(db.String(255), nullable=False, index=True)
    municipio_id = db.Column(db.Integer, db.ForeignKey('municipios.id'), nullable=False)
    direccion    = db.Column(db.String(255))
    telefono     = db.Column(db.String(50))
    contacto     = db.Column(db.String(255))
    activo       = db.Column(db.Boolean, default=True)

    municipio = db.relationship('Municipio',     back_populates='instituciones')
    registros = db.relationship('RecoleccionDato', back_populates='institucion', lazy=True)

    def to_dict(self):
        return {
            'id':              self.id,
            'nombre':          self.nombre,
            'municipio_id':    self.municipio_id,
            'municipio_nombre': self.municipio.nombre if self.municipio else None,
            'direccion':       self.direccion,
            'telefono':        self.telefono,
            'contacto':        self.contacto,
            'activo':          self.activo
        }


class RecoleccionDato(db.Model):
    __tablename__ = 'recoleccion_datos'

    id                    = db.Column(db.Integer, primary_key=True)
    ficha_toma_registro   = db.Column(db.String(50))
    asesor                = db.Column(db.String(255))
    municipio_id          = db.Column(db.Integer, db.ForeignKey('municipios.id'))
    institucion_id        = db.Column(db.Integer, db.ForeignKey('instituciones.id'))
    realizador_nombre     = db.Column(db.String(255))
    realizador_apellidos  = db.Column(db.String(255))
    matricula_documento   = db.Column(db.String(100))
    telefono              = db.Column(db.String(50))
    correo                = db.Column(db.String(255))
    grado                 = db.Column(db.String(50))
    instalacion_educativa = db.Column(db.String(255))
    pista                 = db.Column(db.String(100))
    programa_interes      = db.Column(db.String(255))
    jornada_interes       = db.Column(db.String(100))
    pendido_interes       = db.Column(db.Text)
    asesoria_migradora    = db.Column(db.Text)
    estado                = db.Column(db.String(50), default='pendiente')
    ano_periodo           = db.Column(db.String(50), default='2026-1')
    observacion           = db.Column(db.Text)
    usuario_registro_id   = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    fecha_registro        = db.Column(db.DateTime, default=datetime.utcnow)
    fecha_actualizacion   = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    municipio        = db.relationship('Municipio',   back_populates='registros')
    institucion      = db.relationship('Institucion', back_populates='registros')
    usuario_registro = db.relationship('Usuario',     back_populates='registros')

    def to_dict(self):
        return {
            'id':            self.id,
            'ficha':         self.ficha_toma_registro,
            'asesor':        self.asesor,
            'municipio':     self.municipio.nombre if self.municipio else None,
            'municipio_id':  self.municipio_id,
            'institucion':   self.institucion.nombre if self.institucion else None,
            'institucion_id': self.institucion_id,
            'nombre':        self.realizador_nombre,
            'apellidos':     self.realizador_apellidos,
            'matricula':     self.matricula_documento,
            'telefono':      self.telefono,
            'correo':        self.correo,
            'grado':         self.grado,
            'pista':         self.pista,
            'programa':      self.programa_interes,
            'jornada':       self.jornada_interes,
            'pendido':       self.pendido_interes,
            'asesoria':      self.asesoria_migradora,
            'estado':        self.estado,
            'periodo':       self.ano_periodo,
            'observacion':   self.observacion,
            'fecha':         self.fecha_registro.isoformat() if self.fecha_registro else None,
            'usuario':       self.usuario_registro.nombre if self.usuario_registro else None
        }


class Feria(db.Model):
    __tablename__ = 'ferias'

    id           = db.Column(db.Integer, primary_key=True)
    nombre       = db.Column(db.String(255), nullable=False)
    fecha_inicio = db.Column(db.Date)
    fecha_fin    = db.Column(db.Date)
    ubicacion    = db.Column(db.String(255))
    municipio_id = db.Column(db.Integer, db.ForeignKey('municipios.id'))
    descripcion  = db.Column(db.Text)
    activa       = db.Column(db.Boolean, default=True)

    imagenes     = db.relationship('FeriaImagen', back_populates='feria', lazy=True)
    municipio_rel = db.relationship('Municipio')

    def to_dict(self):
        return {
            'id':             self.id,
            'nombre':         self.nombre,
            'fecha_inicio':   self.fecha_inicio.isoformat() if self.fecha_inicio else None,
            'fecha_fin':      self.fecha_fin.isoformat() if self.fecha_fin else None,
            'ubicacion':      self.ubicacion,
            'municipio':      self.municipio_rel.nombre if self.municipio_rel else None,
            'municipio_id':   self.municipio_id,
            'descripcion':    self.descripcion,
            'activa':         self.activa,
            'total_imagenes': len(self.imagenes) if self.imagenes else 0
        }


class FeriaImagen(db.Model):
    __tablename__ = 'ferias_imagenes'

    id                 = db.Column(db.Integer, primary_key=True)
    feria_id           = db.Column(db.Integer, db.ForeignKey('ferias.id'))
    public_id          = db.Column(db.String(255))
    url                = db.Column(db.String(500))
    descripcion        = db.Column(db.Text)
    usuario_subida_id  = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    fecha_subida       = db.Column(db.DateTime, default=datetime.utcnow)

    feria         = db.relationship('Feria',   back_populates='imagenes')
    usuario_subida = db.relationship('Usuario', back_populates='imagenes_subidas')

    def to_dict(self):
        return {
            'id':        self.id,
            'feria_id':  self.feria_id,
            'public_id': self.public_id,
            'url':       self.url,
            'descripcion': self.descripcion,
            'usuario':   self.usuario_subida.nombre if self.usuario_subida else None,
            'fecha':     self.fecha_subida.isoformat() if self.fecha_subida else None
        }


class ArchivoImportado(db.Model):
    __tablename__ = 'archivos_importados'

    id                   = db.Column(db.Integer, primary_key=True)
    nombre_archivo       = db.Column(db.String(255))
    tipo                 = db.Column(db.String(50))
    url                  = db.Column(db.String(500))
    usuario_importo_id   = db.Column(db.Integer, db.ForeignKey('usuarios.id'))
    fecha_importacion    = db.Column(db.DateTime, default=datetime.utcnow)
    registros_procesados = db.Column(db.Integer, default=0)
    estado               = db.Column(db.String(50), default='completado')
    datos_metadata       = db.Column(db.JSON, default={})

    usuario_importo = db.relationship('Usuario', back_populates='importaciones')

    def to_dict(self):
        return {
            'id':                   self.id,
            'nombre_archivo':       self.nombre_archivo,
            'tipo':                 self.tipo,
            'url':                  self.url,
            'usuario':              self.usuario_importo.nombre if self.usuario_importo else None,
            'fecha_importacion':    self.fecha_importacion.isoformat() if self.fecha_importacion else None,
            'registros_procesados': self.registros_procesados,
            'estado':               self.estado,
            'datos_metadata':       self.datos_metadata
        }


# ============================================================================
# MODELO WHATSAPP
# ============================================================================

class MensajeWhatsApp(db.Model):
    __tablename__ = 'mensajes_whatsapp'

    id             = db.Column(db.Integer, primary_key=True)
    asesora        = db.Column(db.String(50),  nullable=False, index=True)
    direccion      = db.Column(db.String(10),  default='entrante')
    remitente      = db.Column(db.String(100))
    destinatario   = db.Column(db.String(100))
    tipo           = db.Column(db.String(20),  default='text')
    contenido      = db.Column(db.Text)
    media_url      = db.Column(db.Text)
    msg_id         = db.Column(db.String(120), unique=True, index=True)
    timestamp      = db.Column(db.DateTime,    default=datetime.utcnow, index=True)
    fecha_registro = db.Column(db.DateTime,    default=datetime.utcnow)

    def to_dict(self):
        return {
            'id':           self.id,
            'asesora':      self.asesora,
            'direccion':    self.direccion,
            'remitente':    limpiar_numero_contacto(self.remitente),
            'destinatario': limpiar_numero_contacto(self.destinatario),
            'tipo':         self.tipo,
            'contenido':    self.contenido if self.contenido else '(sin contenido)',
            'media_url':    self.media_url,
            'msg_id':       self.msg_id,
            'timestamp':    self.timestamp.isoformat() if self.timestamp else None,
        }


# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================

def allowed_file(filename, file_types=None):
    if '.' not in filename:
        return False
    extension = filename.rsplit('.', 1)[1].lower()
    if file_types == 'image':
        return extension in app.config['ALLOWED_IMAGES']
    return extension in app.config['ALLOWED_EXTENSIONS']


def validate_email_format(email):
    try:
        valid = validate_email(email)
        return valid.email
    except EmailNotValidError:
        return None


def generar_excel_recoleccion(filtros=None):
    query = RecoleccionDato.query
    if filtros:
        if filtros.get('municipio_id'):
            query = query.filter_by(municipio_id=filtros['municipio_id'])
        if filtros.get('ano_periodo'):
            query = query.filter_by(ano_periodo=filtros['ano_periodo'])
        if filtros.get('estado'):
            query = query.filter_by(estado=filtros['estado'])
        if filtros.get('fecha_inicio') and filtros.get('fecha_fin'):
            query = query.filter(
                RecoleccionDato.fecha_registro.between(filtros['fecha_inicio'], filtros['fecha_fin'])
            )

    datos = query.order_by(RecoleccionDato.fecha_registro.desc()).all()
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Recolección 2026-1"

    headers = ['ID', 'Ficha', 'Asesor', 'Municipio', 'Institución', 'Nombre',
               'Apellidos', 'Documento', 'Teléfono', 'Email', 'Grado', 'Programa',
               'Jornada', 'Estado', 'Periodo', 'Observación', 'Fecha Registro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for row, d in enumerate(datos, 2):
        ws.cell(row=row, column=1,  value=d.id)
        ws.cell(row=row, column=2,  value=d.ficha_toma_registro)
        ws.cell(row=row, column=3,  value=d.asesor)
        ws.cell(row=row, column=4,  value=d.municipio.nombre if d.municipio else '')
        ws.cell(row=row, column=5,  value=d.institucion.nombre if d.institucion else '')
        ws.cell(row=row, column=6,  value=d.realizador_nombre)
        ws.cell(row=row, column=7,  value=d.realizador_apellidos)
        ws.cell(row=row, column=8,  value=d.matricula_documento)
        ws.cell(row=row, column=9,  value=d.telefono)
        ws.cell(row=row, column=10, value=d.correo)
        ws.cell(row=row, column=11, value=d.grado)
        ws.cell(row=row, column=12, value=d.programa_interes)
        ws.cell(row=row, column=13, value=d.jornada_interes)
        ws.cell(row=row, column=14, value=d.estado)
        ws.cell(row=row, column=15, value=d.ano_periodo)
        ws.cell(row=row, column=16, value=d.observacion)
        ws.cell(row=row, column=17, value=d.fecha_registro.strftime('%Y-%m-%d %H:%M') if d.fecha_registro else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(output)
    output.seek(0)
    return output


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(Usuario, int(user_id))


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin():
            flash('Acceso denegado. Se requieren permisos de administrador.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function


# ============================================================================
# RUTAS DE AUTENTICACIÓN
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    # Si es POST, proceso el login por email (sistema tradicional)
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        if not email:
            flash('Por favor ingresa tu correo', 'warning')
            return redirect(url_for('login'))

        usuario = Usuario.query.filter_by(email=email).first()
        if usuario and usuario.activo:
            login_user(usuario, remember=True)
            session.permanent = True
            session['user_id'] = usuario.id
            session['user_email'] = usuario.email
            session['user_name'] = usuario.nombre
            session['user_rol'] = usuario.rol
            session['auth_provider'] = 'email'
            session.modified = True
            usuario.ultimo_acceso = datetime.utcnow()
            db.session.commit()
            flash(f'¡Bienvenido {usuario.nombre}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Credenciales inválidas. Verifica tu correo.', 'danger')
            return redirect(url_for('login'))

    # GET: Mostrar página de login con URL de Microsoft si está configurado
    auth_url = None
    if msal_app:
        try:
            # Generar URL de autorización de Microsoft
            auth_url = msal_app.get_authorization_request_url(
                scopes=AZURE_SCOPE,
                redirect_uri=url_for('microsoft_callback', _external=True),
                state=secrets.token_urlsafe(32)
            )
            print(f"🔵 URL de Microsoft generada: {auth_url[:100]}...")
        except Exception as e:
            print(f"❌ Error generando URL de Microsoft: {e}")
            flash('Error al configurar login con Microsoft', 'warning')

    return render_template('login.html', auth_url=auth_url, now=datetime.now)


# ============================================================================
# AUTENTICACIÓN CON MICROSOFT AZURE
# ============================================================================

@app.route('/auth/microsoft/callback')
def microsoft_callback():
    """Callback de Microsoft Azure AD después del login"""
    code = request.args.get('code')
    error = request.args.get('error')
    error_description = request.args.get('error_description')
    
    # Verificar errores
    if error:
        flash(f'Error en autenticación con Microsoft: {error_description or error}', 'danger')
        return redirect(url_for('login'))
    
    if not code:
        flash('No se recibió código de autorización', 'danger')
        return redirect(url_for('login'))
    
    if not msal_app:
        flash('Microsoft Azure no está configurado correctamente', 'danger')
        return redirect(url_for('login'))
    
    try:
        # Intercambiar código por token
        result = msal_app.acquire_token_by_authorization_code(
            code=code,
            scopes=AZURE_SCOPE,
            redirect_uri=url_for('microsoft_callback', _external=True)
        )
        
        if 'error' in result:
            flash(f'Error al obtener token: {result.get("error_description", "Error desconocido")}', 'danger')
            return redirect(url_for('login'))
        
        # Obtener información del usuario desde el token ID
        id_token = result.get('id_token_claims', {})
        
        # Extraer datos del usuario
        user_email = id_token.get('email') or id_token.get('preferred_username') or id_token.get('unique_name')
        user_name = id_token.get('name', '')
        azure_id = id_token.get('oid') or id_token.get('sub')
        first_name = id_token.get('given_name', '')
        last_name = id_token.get('family_name', '')
        
        if not user_email:
            flash('No se pudo obtener el correo del usuario desde Microsoft', 'danger')
            return redirect(url_for('login'))
        
        # Buscar usuario existente
        usuario = Usuario.query.filter_by(email=user_email.lower()).first()
        
        if not usuario:
            # Crear nuevo usuario con datos de Azure
            nombre = first_name or (user_name.split(' ')[0] if user_name else 'Usuario')
            apellido = last_name or (user_name.split(' ')[1] if user_name and len(user_name.split(' ')) > 1 else '')
            
            usuario = Usuario(
                email=user_email.lower(),
                nombre=nombre,
                apellido=apellido,
                azure_id=azure_id,
                rol='usuario',  # Por defecto usuario normal
                activo=True,
                fecha_registro=datetime.utcnow()
            )
            db.session.add(usuario)
            db.session.commit()
            flash('¡Cuenta creada exitosamente con Microsoft!', 'success')
        else:
            # Actualizar datos existentes si es necesario
            if not usuario.azure_id and azure_id:
                usuario.azure_id = azure_id
            if not usuario.nombre and first_name:
                usuario.nombre = first_name
            if not usuario.apellido and last_name:
                usuario.apellido = last_name
            
            # Verificar si la cuenta está activa
            if not usuario.activo:
                flash('Tu cuenta está desactivada. Contacta al administrador.', 'danger')
                return redirect(url_for('login'))
            
            db.session.commit()
            flash(f'Bienvenido de vuelta {usuario.nombre}!', 'success')
        
        # Iniciar sesión
        login_user(usuario, remember=True)
        session.permanent = True
        session['user_id'] = usuario.id
        session['user_email'] = usuario.email
        session['user_name'] = usuario.nombre
        session['user_rol'] = usuario.rol
        session['auth_provider'] = 'microsoft'
        
        # Actualizar último acceso
        usuario.ultimo_acceso = datetime.utcnow()
        db.session.commit()
        
        return redirect(url_for('dashboard'))
        
    except Exception as e:
        app.logger.error(f"Error en callback Microsoft: {str(e)}")
        flash('Error al procesar la autenticación con Microsoft. Intenta nuevamente.', 'danger')
        return redirect(url_for('login'))


@app.route('/logout')
@login_required
def logout():
    # Obtener información de la sesión antes de cerrarla
    auth_provider = session.get('auth_provider')
    
    # Cerrar sesión local
    logout_user()
    session.clear()
    
    flash('Sesión cerrada exitosamente', 'info')
    
    # Si inició sesión con Microsoft, redirigir a logout de Microsoft también
    if auth_provider == 'microsoft' and msal_app:
        microsoft_logout_url = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/logout?post_logout_redirect_uri={url_for('login', _external=True)}"
        return redirect(microsoft_logout_url)
    
    return redirect(url_for('login'))


# ============================================================================
# RUTAS PRINCIPALES
# ============================================================================

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    try:
        total_registros     = RecoleccionDato.query.filter_by(ano_periodo='2026-1').count()
        total_municipios    = db.session.query(func.count(db.distinct(RecoleccionDato.municipio_id))).filter_by(ano_periodo='2026-1').scalar() or 0
        total_instituciones = db.session.query(func.count(db.distinct(RecoleccionDato.institucion_id))).filter_by(ano_periodo='2026-1').scalar() or 0
        completados         = RecoleccionDato.query.filter_by(ano_periodo='2026-1', estado='completado').count()
        pendientes          = RecoleccionDato.query.filter_by(ano_periodo='2026-1', estado='pendiente').count()
        registros_recientes = RecoleccionDato.query.order_by(RecoleccionDato.fecha_registro.desc()).limit(10).all()

        return render_template('dashboard.html',
                               total_registros=total_registros,
                               total_municipios=total_municipios,
                               total_instituciones=total_instituciones,
                               completados=completados,
                               pendientes=pendientes,
                               registros_recientes=registros_recientes,
                               usuario=current_user,
                               now=datetime.now)
    except Exception:
        app.logger.error("Error en dashboard")
        flash('Error al cargar el dashboard', 'error')
        return render_template('dashboard.html', usuario=current_user, now=datetime.now)


# ============================================================================
# RUTAS DE RECOLECCIÓN DE DATOS
# ============================================================================

@app.route('/recoleccion')
@login_required
def recoleccion():
    # Obtener municipios (filtrados por activos)
    municipios = Municipio.query.filter_by(activo=True).order_by(Municipio.nombre).all()

    # Obtener instituciones (para el select)
    instituciones = Institucion.query.options(db.joinedload(Institucion.municipio)).filter_by(activo=True).all()

    municipio_id = request.args.get('municipio', type=int)
    periodo = request.args.get('periodo', '2026-1')

    query = RecoleccionDato.query.filter_by(ano_periodo=periodo)
    if municipio_id:
        query = query.filter_by(municipio_id=municipio_id)

    registros = query.order_by(RecoleccionDato.fecha_registro.desc()).limit(100).all()

    return render_template('recoleccion.html',
                           registros=registros,
                           municipios=municipios,
                           instituciones=instituciones,
                           periodo_actual=periodo,
                           now=datetime.now)


@app.route('/api/recoleccion', methods=['POST'])
@login_required
def api_crear_recoleccion():
    try:
        data = request.get_json()

        if data.get('correo'):
            email_validado = validate_email_format(data['correo'])
            if not email_validado:
                return jsonify({'error': 'Email inválido'}), 400
            data['correo'] = email_validado

        def parse_int(val):
            if val in ('', None):
                return None
            try:
                return int(val)
            except Exception:
                return None

        nuevo_registro = RecoleccionDato(
            ficha_toma_registro  = data.get('ficha_toma_registro'),
            asesor               = data.get('asesor'),
            municipio_id         = parse_int(data.get('municipio_id')),
            institucion_id       = parse_int(data.get('institucion_id')),
            realizador_nombre    = data.get('realizador_nombre'),
            realizador_apellidos = data.get('realizador_apellidos'),
            matricula_documento  = data.get('matricula_documento'),
            telefono             = data.get('telefono'),
            correo               = data.get('correo'),
            grado                = data.get('grado'),
            instalacion_educativa= data.get('instalacion_educativa'),
            pista                = data.get('pista'),
            programa_interes     = data.get('programa_interes'),
            jornada_interes      = data.get('jornada_interes'),
            pendido_interes      = data.get('pendido_interes'),
            asesoria_migradora   = data.get('asesoria_migradora'),
            estado               = data.get('estado', 'pendiente'),
            ano_periodo          = data.get('ano_periodo', '2026-1'),
            observacion          = data.get('observacion'),
            usuario_registro_id  = current_user.id
        )

        db.session.add(nuevo_registro)
        db.session.commit()
        return jsonify({'message': 'Registro creado exitosamente', 'data': nuevo_registro.to_dict()}), 201

    except Exception:
        db.session.rollback()
        app.logger.error("Error al crear registro")
        return jsonify({'error': 'Error al crear el registro'}), 500


@app.route('/api/recoleccion/<int:id>', methods=['PUT', 'GET'])
@login_required
def api_obtener_recoleccion(id):
    try:
        registro = RecoleccionDato.query.get_or_404(id)

        if request.method == 'GET':
            return jsonify(registro.to_dict())

        data = request.get_json()

        if data.get('correo'):
            email_validado = validate_email_format(data['correo'])
            if not email_validado:
                return jsonify({'error': 'Email inválido'}), 400
            data['correo'] = email_validado

        for field in ('municipio_id', 'institucion_id'):
            if field in data:
                if data[field] in ('', None):
                    data[field] = None
                else:
                    try:
                        data[field] = int(data[field])
                    except Exception:
                        data[field] = None

        for key, value in data.items():
            if hasattr(registro, key):
                setattr(registro, key, value)

        registro.fecha_actualizacion = datetime.utcnow()
        db.session.commit()
        return jsonify({'message': 'Registro actualizado exitosamente', 'data': registro.to_dict()})

    except Exception:
        db.session.rollback()
        return jsonify({'error': 'Error al actualizar el registro'}), 500


# ============================================================================
# RUTAS DE FERIAS
# ============================================================================

@app.route('/ferias')
@login_required
def ferias():
    ferias_list = Feria.query.filter_by(activa=True).all()
    municipios  = Municipio.query.filter_by(activo=True).all()
    return render_template('ferias.html',
                           ferias=ferias_list,
                           municipios=municipios,
                           now=datetime.now,
                           cloudinary_configurado=CLOUDINARY_CONFIGURED)


@app.route('/api/ferias', methods=['POST'])
@login_required
def api_crear_feria():
    try:
        data = request.get_json()
        fecha_inicio = None
        fecha_fin    = None

        if data.get('fecha_inicio'):
            try:
                fecha_inicio = datetime.strptime(data['fecha_inicio'], '%Y-%m-%d').date()
            except Exception:
                pass

        if data.get('fecha_fin'):
            try:
                fecha_fin = datetime.strptime(data['fecha_fin'], '%Y-%m-%d').date()
            except Exception:
                pass

        nueva_feria = Feria(
            nombre=data.get('nombre'),
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            ubicacion=data.get('ubicacion'),
            municipio_id=data.get('municipio_id'),
            descripcion=data.get('descripcion'),
            activa=True
        )
        db.session.add(nueva_feria)
        db.session.commit()
        return jsonify({'message': 'Feria creada exitosamente', 'data': nueva_feria.to_dict()}), 201

    except Exception:
        db.session.rollback()
        return jsonify({'error': 'Error al crear la feria'}), 500


@app.route('/api/ferias/<int:feria_id>/imagenes', methods=['POST'])
@login_required
def api_subir_imagen_feria(feria_id):
    try:
        # Verificar que la feria existe
        Feria.query.get_or_404(feria_id)

        if 'images' not in request.files:
            return jsonify({'error': 'No se enviaron imágenes'}), 400

        files             = request.files.getlist('images')
        imagenes_subidas  = []
        errores           = []

        for file in files:
            if file and allowed_file(file.filename, 'image'):
                try:
                    filename    = secure_filename(file.filename)
                    timestamp   = datetime.now().strftime('%Y%m%d_%H%M%S')
                    random_hex  = secrets.token_hex(4)
                    url = public_id = ""

                    if CLOUDINARY_CONFIGURED:
                        result    = cloudinary.uploader.upload(
                            file,
                            folder=f'uniagraria/ferias/{feria_id}',
                            public_id=f"{feria_id}_{timestamp}_{random_hex}",
                            transformation=[{'width': 1200, 'height': 800, 'crop': 'limit'}, {'quality': 'auto'}]
                        )
                        url       = result['secure_url']
                        public_id = result['public_id']
                    else:
                        safe_filename = f"{timestamp}_{random_hex}_{filename}"
                        file_path     = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
                        file.save(file_path)
                        url       = url_for('static', filename=f'uploads/{safe_filename}', _external=True)
                        public_id = safe_filename

                    imagen = FeriaImagen(
                        feria_id=feria_id, public_id=public_id, url=url,
                        descripcion=request.form.get('descripcion', ''),
                        usuario_subida_id=current_user.id, fecha_subida=datetime.utcnow()
                    )
                    db.session.add(imagen)
                    db.session.flush()
                    imagenes_subidas.append({'url': url, 'public_id': public_id, 'id': imagen.id})

                except Exception:
                    errores.append({'filename': file.filename, 'error': 'Error al subir'})
                    db.session.rollback()
            else:
                errores.append({'filename': file.filename, 'error': 'Tipo de archivo no permitido'})

        if imagenes_subidas:
            db.session.commit()

        verificacion = FeriaImagen.query.filter_by(feria_id=feria_id).count()
        return jsonify({
            'message': f'{len(imagenes_subidas)} imágenes subidas exitosamente',
            'imagenes': imagenes_subidas, 'errores': errores,
            'total_en_bd': verificacion, 'cloudinary_configurado': CLOUDINARY_CONFIGURED
        }), 200

    except Exception:
        db.session.rollback()
        return jsonify({'error': 'Error al subir las imágenes'}), 500


@app.route('/api/ferias/<int:feria_id>/imagenes', methods=['GET'])
@login_required
def api_obtener_imagenes_feria(feria_id):
    try:
        # Verificar que la feria existe
        Feria.query.get_or_404(feria_id)
        imagenes = FeriaImagen.query.filter_by(feria_id=feria_id).order_by(FeriaImagen.fecha_subida.desc()).all()
        return jsonify({'imagenes': [i.to_dict() for i in imagenes], 'total': len(imagenes)})
    except Exception:
        return jsonify({'error': 'Error al obtener imágenes'}), 500


@app.route('/api/ferias/imagenes/<path:public_id>', methods=['DELETE'])
@login_required
def api_eliminar_imagen(public_id):
    try:
        imagen = FeriaImagen.query.filter_by(public_id=public_id).first_or_404()
        if CLOUDINARY_CONFIGURED:
            try:
                cloudinary.uploader.destroy(public_id)
            except Exception:
                pass
        else:
            try:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], public_id)
                if os.path.exists(filepath):
                    os.remove(filepath)
            except Exception:
                pass
        db.session.delete(imagen)
        db.session.commit()
        return jsonify({'message': 'Imagen eliminada exitosamente'})
    except Exception:
        db.session.rollback()
        return jsonify({'error': 'Error al eliminar la imagen'}), 500


# ============================================================================
# RUTAS DE IMPORTACIÓN
# ============================================================================

@app.route('/importacion')
@login_required
@admin_required
def importacion():
    importaciones = ArchivoImportado.query.order_by(ArchivoImportado.fecha_importacion.desc()).limit(20).all()
    return render_template('importacion.html', importaciones=importaciones, now=datetime.now)


@app.route('/api/importacion/excel', methods=['POST'])
@login_required
@admin_required
@limiter.limit("10 per minute")
def api_importar_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se envió ningún archivo'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nombre de archivo vacío'}), 400
        if not allowed_file(file.filename):
            return jsonify({'error': 'Tipo de archivo no permitido'}), 400

        filename = secure_filename(file.filename)

        try:
            wb      = load_workbook(file)
            ws      = wb.active
            headers = [cell.value for cell in ws[1]]
            data    = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    data.append(dict(zip(headers, row)))
        except Exception:
            return jsonify({'error': 'Error al leer el archivo Excel'}), 400

        if not data:
            return jsonify({'error': 'El archivo está vacío'}), 400

        registros_procesados = 0
        errores = []

        for idx, row in enumerate(data, 2):
            try:
                municipio_nombre = str(row.get('Municipio', row.get('municipio', ''))).strip()
                municipio = None
                if municipio_nombre:
                    municipio = Municipio.query.filter(func.lower(Municipio.nombre) == func.lower(municipio_nombre)).first()
                    if not municipio:
                        municipio = Municipio(nombre=municipio_nombre)
                        db.session.add(municipio)
                        db.session.flush()

                institucion_nombre = str(row.get('Instalacion_Educativa', row.get('instalacion_educativa', ''))).strip()
                institucion = None
                if institucion_nombre and municipio:
                    institucion = Institucion.query.filter(
                        func.lower(Institucion.nombre) == func.lower(institucion_nombre),
                        Institucion.municipio_id == municipio.id
                    ).first()
                    if not institucion:
                        institucion = Institucion(nombre=institucion_nombre, municipio_id=municipio.id)
                        db.session.add(institucion)
                        db.session.flush()

                registro = RecoleccionDato(
                    ficha_toma_registro  = str(row.get('Ficha_toma_registro',   row.get('ficha', '')))[:50],
                    asesor               = str(row.get('Asesor',                row.get('asesor', '')))[:255],
                    municipio_id         = municipio.id if municipio else None,
                    institucion_id       = institucion.id if institucion else None,
                    realizador_nombre    = str(row.get('Realizador_Nombre',     row.get('nombre', '')))[:255],
                    realizador_apellidos = str(row.get('Realizador_Apellidos',  row.get('apellidos', '')))[:255],
                    matricula_documento  = str(row.get('Matricula_Documento',   row.get('matricula', '')))[:100],
                    telefono             = str(row.get('Telefono',              row.get('telefono', '')))[:50],
                    correo               = str(row.get('Correo',               row.get('correo', '')))[:255],
                    grado                = str(row.get('Grado',                row.get('grado', '')))[:50],
                    instalacion_educativa= str(row.get('Instalacion_Educativa', row.get('instalacion_educativa', '')))[:255],
                    pista                = str(row.get('Pista',                row.get('pista', '')))[:100],
                    programa_interes     = str(row.get('Programa_Interes',     row.get('programa_interes', '')))[:255],
                    jornada_interes      = str(row.get('Jornada_Interes',      row.get('jornada_interes', '')))[:100],
                    pendido_interes      = str(row.get('Pendido_Interes',      row.get('pendido_interes', ''))),
                    asesoria_migradora   = str(row.get('Asesoria_Migradora',   row.get('asesoria_migradora', ''))),
                    estado               = str(row.get('Estado',               row.get('estado', 'pendiente'))),
                    ano_periodo          = str(row.get('Ano_periodo',          row.get('ano_periodo', '2026-1'))),
                    observacion          = str(row.get('Observacion',          row.get('observacion', ''))),
                    usuario_registro_id  = current_user.id
                )
                db.session.add(registro)
                registros_procesados += 1
                if registros_procesados % 50 == 0:
                    db.session.commit()

            except Exception:
                db.session.rollback()
                errores.append({'fila': idx, 'error': 'Error en fila'})

        db.session.commit()

        try:
            archivo_importado = ArchivoImportado(
                nombre_archivo=filename, tipo='excel', url='',
                usuario_importo_id=current_user.id,
                registros_procesados=registros_procesados,
                estado='completado',
                datos_metadata={'total_filas': len(data), 'procesados': registros_procesados, 'errores': len(errores)}
            )
            db.session.add(archivo_importado)
            db.session.commit()
        except Exception:
            app.logger.error("Error al guardar metadata")

        return jsonify({
            'message': f'Importación completada: {registros_procesados} registros procesados',
            'registros_procesados': registros_procesados,
            'errores': len(errores),
            'detalles_errores': errores[:10]
        })

    except Exception:
        db.session.rollback()
        return jsonify({'error': 'Error interno en el servidor'}), 500


# ============================================================================
# RUTAS DE REPORTES
# ============================================================================

@app.route('/reportes')
@login_required
def reportes():
    municipios = Municipio.query.filter_by(activo=True).all()
    return render_template('reportes.html', municipios=municipios, now=datetime.now)


@app.route('/api/reportes/general')
@login_required
def api_reporte_general():
    try:
        periodo      = request.args.get('periodo', '2026-1')
        municipio_id = request.args.get('municipio_id', type=int)
        estado       = request.args.get('estado')
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin    = request.args.get('fecha_fin')

        query = RecoleccionDato.query.filter_by(ano_periodo=periodo)
        if municipio_id:
            query = query.filter_by(municipio_id=municipio_id)
        if estado:
            query = query.filter_by(estado=estado)
        if fecha_inicio and fecha_fin:
            query = query.filter(RecoleccionDato.fecha_registro.between(fecha_inicio, fecha_fin))

        registros = query.all()
        stats = {
            'total':       len(registros),
            'completados': sum(1 for r in registros if r.estado == 'completado'),
            'pendientes':  sum(1 for r in registros if r.estado == 'pendiente'),
            'por_grado': {}, 'por_programa': {}, 'por_municipio': {}
        }
        for r in registros:
            if r.grado:
                stats['por_grado'][r.grado] = stats['por_grado'].get(r.grado, 0) + 1
            if r.programa_interes:
                stats['por_programa'][r.programa_interes] = stats['por_programa'].get(r.programa_interes, 0) + 1
            if r.municipio:
                stats['por_municipio'][r.municipio.nombre] = stats['por_municipio'].get(r.municipio.nombre, 0) + 1

        return jsonify({'periodo': periodo, 'estadisticas': stats, 'registros': [r.to_dict() for r in registros[:100]]})

    except Exception:
        return jsonify({'error': 'Error al generar reporte'}), 500


@app.route('/api/reportes/exportar/excel')
@login_required
def api_exportar_excel():
    try:
        filtros = {
            'municipio_id': request.args.get('municipio_id', type=int),
            'ano_periodo':  request.args.get('periodo', '2026-1'),
            'estado':       request.args.get('estado'),
            'fecha_inicio': request.args.get('fecha_inicio'),
            'fecha_fin':    request.args.get('fecha_fin')
        }
        output = generar_excel_recoleccion(filtros)
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=f'reporte_uniagraria_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
    except Exception:
        flash('Error al generar el archivo Excel', 'error')
        return redirect(url_for('reportes'))


@app.route('/api/reportes/exportar/pdf')
@login_required
def api_exportar_pdf():
    try:
        periodo   = request.args.get('periodo', '2026-1')
        registros = RecoleccionDato.query.filter_by(ano_periodo=periodo).all()
        buffer    = BytesIO()
        doc       = SimpleDocTemplate(buffer, pagesize=A4)
        elements  = []
        styles    = getSampleStyleSheet()

        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                     fontSize=24, textColor=colors.HexColor('#2E7D32'),
                                     spaceAfter=30, alignment=1)
        info_style  = ParagraphStyle('InfoStyle',  parent=styles['Normal'], fontSize=12, spaceAfter=10)

        elements.append(Paragraph(f'UNIAGRARIA - RECOLECCIÓN FACATATIVÁ {periodo}', title_style))
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f'Fecha: {datetime.now().strftime("%d/%m/%Y %H:%M")}', info_style))
        elements.append(Paragraph(f'Total registros: {len(registros)}', info_style))
        elements.append(Paragraph(f'Generado por: {current_user.nombre}', info_style))
        elements.append(Spacer(1, 30))

        data = [['ID', 'Nombre', 'Municipio', 'Institución', 'Grado', 'Estado']]
        for r in registros[:50]:
            data.append([str(r.id),
                         f"{r.realizador_nombre} {r.realizador_apellidos or ''}",
                         r.municipio.nombre if r.municipio else 'N/A',
                         r.institucion.nombre[:30] if r.institucion else 'N/A',
                         r.grado or 'N/A', r.estado])

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN',      (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID',       (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        return send_file(buffer, mimetype='application/pdf', as_attachment=True,
                         download_name=f'reporte_uniagraria_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf')
    except Exception:
        flash('Error al generar el PDF', 'error')
        return redirect(url_for('reportes'))


# ============================================================================
# RUTAS DE ARCHIVOS
# ============================================================================

@app.route('/archivos')
@login_required
def archivos():
    importaciones = ArchivoImportado.query.order_by(ArchivoImportado.fecha_importacion.desc()).limit(30).all()
    return render_template('archivos.html', importaciones=importaciones, now=datetime.now)


@app.route('/api/archivos/comprimir-todo')
@login_required
@admin_required
def api_comprimir_todo():
    try:
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            registros    = RecoleccionDato.query.filter_by(ano_periodo='2026-1').all()
            ferias_list  = Feria.query.all()
            imagenes     = FeriaImagen.query.all()

            zf.writestr('recoleccion_datos_2026-1.json', json.dumps([r.to_dict() for r in registros], indent=2, default=str))
            zf.writestr('ferias.json',           json.dumps([f.to_dict() for f in ferias_list], indent=2, default=str))
            zf.writestr('ferias_imagenes.json',  json.dumps([i.to_dict() for i in imagenes], indent=2, default=str))
            zf.writestr('recoleccion_datos_2026-1.xlsx', generar_excel_recoleccion().getvalue())
            zf.writestr('README.txt', f"UNIAGRARIA backup {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
                                      f"Registros: {len(registros)} | Ferias: {len(ferias_list)} | Imágenes: {len(imagenes)}")

        zip_buffer.seek(0)
        return send_file(zip_buffer, mimetype='application/zip', as_attachment=True,
                         download_name=f'uniagraria_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip')
    except Exception:
        flash('Error al generar el archivo ZIP', 'error')
        return redirect(url_for('archivos'))


# ============================================================================
# RUTAS DE ADMINISTRACIÓN
# ============================================================================

@app.route('/admin/usuarios')
@login_required
@admin_required
def admin_usuarios():
    usuarios = Usuario.query.order_by(Usuario.fecha_registro.desc()).all()
    return render_template('admin_usuarios.html', usuarios=usuarios, now=datetime.now)


@app.route('/api/admin/usuarios/<int:user_id>/rol', methods=['PUT'])
@login_required
@admin_required
def api_cambiar_rol(user_id):
    try:
        usuario  = Usuario.query.get_or_404(user_id)
        data     = request.get_json()
        nuevo_rol = data.get('rol')
        if nuevo_rol in ['admin', 'usuario']:
            usuario.rol = nuevo_rol
            db.session.commit()
            return jsonify({'message': 'Rol actualizado exitosamente'})
        return jsonify({'error': 'Rol inválido'}), 400
    except Exception:
        db.session.rollback()
        return jsonify({'error': 'Error al cambiar rol'}), 500


@app.route('/api/importacion/detalles/<int:id>', methods=['GET'])
@login_required
def api_detalles_importacion(id):
    try:
        importacion = ArchivoImportado.query.get_or_404(id)
        return jsonify({
            'id':                   importacion.id,
            'nombre_archivo':       importacion.nombre_archivo,
            'fecha_importacion':    importacion.fecha_importacion.isoformat() if importacion.fecha_importacion else None,
            'usuario':              importacion.usuario_importo.nombre if importacion.usuario_importo else None,
            'registros_procesados': importacion.registros_procesados,
            'estado':               importacion.estado,
            'tipo':                 importacion.tipo,
            'url':                  importacion.url,
            'datos_metadata':       importacion.datos_metadata
        })
    except Exception:
        return jsonify({'error': 'Error al obtener detalles'}), 500


# ============================================================================
# MÓDULO WHATSAPP - RUTAS
# ============================================================================

# Página web de WhatsApp (requiere login)
@app.route('/whatsapp')
@login_required
def whatsapp():
    asesoras = list(WA_NOMBRES.items())
    return render_template('whatsapp.html', asesoras=asesoras, now=datetime.now)


# ENDPOINTS PÚBLICOS PARA LA API DE WHATSAPP (NO requieren autenticación)
@app.route('/api/whatsapp/mensaje', methods=['POST', 'OPTIONS'])
def api_whatsapp_recibir():
    """Recibe mensajes del servicio Node.js - SIN filtrar @lid"""

    if request.method == 'OPTIONS':
        response = make_response()
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type, X-WA-Token')
        response.headers.add('Access-Control-Allow-Methods', 'POST')
        return response

    data = request.get_json(silent=True)
    if not data:
        return jsonify({'error': 'JSON inválido'}), 400

    remitente = data.get('remitente', '')
    destinatario = data.get('destinatario', '')

    # SOLO filtrar mensajes de sistema y grupos (NO filtrar @lid)
    if remitente == 'status@broadcast' or destinatario == 'status@broadcast':
        return jsonify({'status': 'ignorado', 'reason': 'mensaje de sistema'}), 200

    if '@g.us' in remitente or '@g.us' in destinatario:
        return jsonify({'status': 'ignorado', 'reason': 'mensaje de grupo'}), 200

    # IMPORTANTE: NO filtramos @lid - son contactos válidos de WhatsApp

    # Evitar duplicados
    if MensajeWhatsApp.query.filter_by(msg_id=data.get('msg_id')).first():
        return jsonify({'status': 'duplicado'}), 200

    try:
        msg = MensajeWhatsApp(
            asesora      = data.get('asesora', 'desconocida'),
            direccion    = data.get('direccion', 'entrante'),
            remitente    = remitente,
            destinatario = destinatario,
            tipo         = data.get('tipo', 'text'),
            contenido    = data.get('contenido', '') or '(sin contenido)',
            media_url    = data.get('media_url'),
            msg_id       = data.get('msg_id'),
            timestamp    = datetime.fromisoformat(data['timestamp']) if data.get('timestamp') else datetime.utcnow(),
        )
        db.session.add(msg)
        db.session.commit()
        print(f"✅ Mensaje guardado: ID={msg.id} | Asesora={msg.asesora} | Remitente={msg.remitente}")
        return jsonify({'status': 'ok', 'id': msg.id}), 201
    except Exception:
        db.session.rollback()
        app.logger.error('Error guardando WA mensaje')
        return jsonify({'error': 'Error interno'}), 500


@app.route('/api/whatsapp/stats', methods=['GET'])
def api_whatsapp_stats():
    """Estadísticas de WhatsApp - Endpoint PÚBLICO"""
    stats = {}
    for key, nombre in WA_NOMBRES.items():
        total = MensajeWhatsApp.query.filter_by(asesora=key).count()
        hoy = MensajeWhatsApp.query.filter(
            MensajeWhatsApp.asesora == key,
            MensajeWhatsApp.timestamp >= datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        ).count()
        ultimo = MensajeWhatsApp.query.filter_by(asesora=key).order_by(MensajeWhatsApp.timestamp.desc()).first()
        stats[key] = {
            'nombre': nombre,
            'total': total,
            'hoy': hoy,
            'ultimo_mensaje': ultimo.timestamp.isoformat() if ultimo else None,
        }
    return jsonify(stats)


@app.route('/api/whatsapp/historial/<asesora>', methods=['GET'])
def api_whatsapp_historial(asesora):
    """Historial de mensajes - Endpoint PÚBLICO"""
    page = request.args.get('page', 1, type=int)
    por_pag = request.args.get('por_pagina', 50, type=int)

    q = MensajeWhatsApp.query.filter_by(asesora=asesora)
    pag = q.order_by(MensajeWhatsApp.timestamp.desc()).paginate(page=page, per_page=por_pag)

    return jsonify({
        'total': pag.total,
        'pagina': pag.page,
        'paginas': pag.pages,
        'mensajes': [m.to_dict() for m in pag.items],
    })


@app.route('/api/whatsapp/conversaciones/<asesora>', methods=['GET'])
def api_whatsapp_conversaciones(asesora):
    """Lista de conversaciones agrupadas por contacto - Endpoint PÚBLICO"""
    try:
        from sqlalchemy import func

        conversaciones = {}

        # Mensajes ENTRANTES
        entrantes = db.session.query(
            MensajeWhatsApp.remitente,
            func.count(MensajeWhatsApp.id).label('total'),
            func.max(MensajeWhatsApp.timestamp).label('ultimo')
        ).filter(
            MensajeWhatsApp.asesora == asesora,
            MensajeWhatsApp.direccion == 'entrante',
            MensajeWhatsApp.remitente.isnot(None),
            MensajeWhatsApp.remitente != '',
            MensajeWhatsApp.remitente != 'status@broadcast',
            ~MensajeWhatsApp.remitente.like('%@g.us')
        ).group_by(MensajeWhatsApp.remitente).all()

        for r in entrantes:
            if r.remitente:
                conversaciones[r.remitente] = {
                    'recibidos': r.total,
                    'enviados': 0,
                    'ultimo': r.ultimo
                }

        # Mensajes SALIENTES
        salientes = db.session.query(
            MensajeWhatsApp.destinatario,
            func.count(MensajeWhatsApp.id).label('total'),
            func.max(MensajeWhatsApp.timestamp).label('ultimo')
        ).filter(
            MensajeWhatsApp.asesora == asesora,
            MensajeWhatsApp.direccion == 'saliente',
            MensajeWhatsApp.destinatario.isnot(None),
            MensajeWhatsApp.destinatario != '',
            MensajeWhatsApp.destinatario != 'status@broadcast',
            ~MensajeWhatsApp.destinatario.like('%@g.us')
        ).group_by(MensajeWhatsApp.destinatario).all()

        for r in salientes:
            if r.destinatario:
                if r.destinatario in conversaciones:
                    conversaciones[r.destinatario]['enviados'] = r.total
                    if r.ultimo and r.ultimo > conversaciones[r.destinatario]['ultimo']:
                        conversaciones[r.destinatario]['ultimo'] = r.ultimo
                else:
                    conversaciones[r.destinatario] = {
                        'recibidos': 0,
                        'enviados': r.total,
                        'ultimo': r.ultimo
                    }

        resultado = []
        for contacto, data in conversaciones.items():
            ultimo_msg = MensajeWhatsApp.query.filter(
                MensajeWhatsApp.asesora == asesora,
                (MensajeWhatsApp.remitente == contacto) | (MensajeWhatsApp.destinatario == contacto)
            ).order_by(MensajeWhatsApp.timestamp.desc()).first()

            resultado.append({
                'contacto': contacto,
                'total_mensajes': data['recibidos'] + data['enviados'],
                'recibidos': data['recibidos'],
                'enviados': data['enviados'],
                'ultimo_mensaje': data['ultimo'].isoformat() if data['ultimo'] else None,
                'ultimo_contenido': ultimo_msg.contenido[:80] if ultimo_msg else '',
                'ultimo_direccion': ultimo_msg.direccion if ultimo_msg else ''
            })

        resultado.sort(key=lambda x: x['ultimo_mensaje'] or '', reverse=True)
        return jsonify(resultado)

    except Exception:
        app.logger.error("Error en conversaciones")
        return jsonify({'error': 'Error interno'}), 500


@app.route('/api/whatsapp/conversacion/<asesora>/<contacto>', methods=['GET'])
def api_whatsapp_conversacion(asesora, contacto):
    """Mensajes con un contacto específico - Endpoint PÚBLICO"""
    try:
        page = request.args.get('page', 1, type=int)
        por_pag = request.args.get('por_pagina', 50, type=int)

        q = MensajeWhatsApp.query.filter(
            MensajeWhatsApp.asesora == asesora,
            (MensajeWhatsApp.remitente == contacto) | (MensajeWhatsApp.destinatario == contacto)
        ).order_by(MensajeWhatsApp.timestamp.asc())

        pag = q.paginate(page=page, per_page=por_pag, error_out=False)

        return jsonify({
            'contacto': contacto,
            'total': pag.total,
            'pagina': pag.page,
            'paginas': pag.pages,
            'mensajes': [m.to_dict() for m in pag.items]
        })

    except Exception:
        app.logger.error("Error en conversacion")
        return jsonify({'error': 'Error interno'}), 500


@app.route('/api/whatsapp/exportar/<asesora>')
@login_required
def api_whatsapp_exportar(asesora):
    mensajes = MensajeWhatsApp.query.filter_by(asesora=asesora).order_by(MensajeWhatsApp.timestamp.asc()).all()

    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = f'Historial {WA_NOMBRES.get(asesora, asesora)}'

    headers = ['Fecha/Hora', 'Dirección', 'Remitente', 'Destinatario', 'Tipo', 'Contenido', 'Media URL']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(start_color='25D366', end_color='25D366', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row, m in enumerate(mensajes, 2):
        ws.cell(row=row, column=1, value=m.timestamp.strftime('%Y-%m-%d %H:%M') if m.timestamp else '')
        ws.cell(row=row, column=2, value='↓ Entrante' if m.direccion == 'entrante' else '↑ Saliente')
        ws.cell(row=row, column=3, value=m.remitente)
        ws.cell(row=row, column=4, value=m.destinatario)
        ws.cell(row=row, column=5, value=m.tipo)
        ws.cell(row=row, column=6, value=m.contenido or '(sin contenido)')
        ws.cell(row=row, column=7, value=m.media_url or '')

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(output)
    output.seek(0)

    nombre = WA_NOMBRES.get(asesora, asesora).replace(' ', '_')
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'historial_wa_{nombre}_{datetime.now().strftime("%Y%m%d")}.xlsx')


# ============================================================================
# MANEJO DE ERRORES
# ============================================================================

@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html', error_code=404, error_message='Página no encontrada', now=datetime.now), 404


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('error.html', error_code=500, error_message='Error interno del servidor', now=datetime.now), 500


@app.errorhandler(429)
def ratelimit_error(error):
    return render_template('error.html', error_code=429, error_message='Demasiadas solicitudes. Por favor espere.', now=datetime.now), 429


# ============================================================================
# INICIALIZACIÓN DE LA BASE DE DATOS
# ============================================================================

def init_database():
    with app.app_context():
        try:
            db.create_all()
            print("✅ Tablas creadas/verificadas en NeonTech")

            municipios_iniciales = ['Facatativá', 'Bogotá', 'Madrid', 'Mosquera', 'Funza', 'El Rosal', 'Subachoque', 'Zipacón']
            for m in municipios_iniciales:
                if not Municipio.query.filter_by(nombre=m).first():
                    db.session.add(Municipio(nombre=m))

            admins_default = [
                {'email': 'admin1@uniagraria.edu.co', 'nombre': 'Administrador Principal',      'rol': 'admin'},
                {'email': 'admin2@uniagraria.edu.co', 'nombre': 'Coordinador de Recolección',   'rol': 'admin'},
                {'email': 'admin3@uniagraria.edu.co', 'nombre': 'Director de Proyectos',        'rol': 'admin'},
                {'email': 'admin4@uniagraria.edu.co', 'nombre': 'Supervisor de Campo',          'rol': 'admin'},
                {'email': 'admin5@uniagraria.edu.co', 'nombre': 'Gestor de Calidad',            'rol': 'admin'},
            ]
            for admin in admins_default:
                if not Usuario.query.filter_by(email=admin['email']).first():
                    db.session.add(Usuario(**admin))

            db.session.commit()
            print("✅ Base de datos inicializada correctamente")
            crear_plantillas_ejemplo()
            return True

        except Exception:
            db.session.rollback()
            print("❌ Error al inicializar base de datos")
            return False


def crear_plantillas_ejemplo():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla"
        headers = ['Ficha_toma_registro', 'Asesor', 'Municipio', 'Instalacion_Educativa',
                   'Realizador_Nombre', 'Realizador_Apellidos', 'Matricula_Documento',
                   'Telefono', 'Correo', 'Grado', 'Programa_Interes', 'Jornada_Interes', 'Estado', 'Observacion']
        ejemplo = ['F001', 'Juan Pérez', 'Facatativá', 'Institución Ejemplo',
                   'María', 'González', '123456', '3001234567',
                   'maria@ejemplo.com', '11°', 'Ingeniería Agronómica', 'Diurna', 'pendiente', 'Registro de ejemplo']

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        for col, valor in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col, value=valor)

        wb.save('static/plantillas/plantilla_recoleccion.xlsx')

        with open('static/plantillas/plantilla_recoleccion.csv', 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerow(ejemplo)

        print("✅ Plantillas de ejemplo creadas")
    except Exception:
        print("⚠️ Error al crear plantillas")


with app.app_context():
    init_database()

# ============================================================================
# PUNTO DE ENTRADA
# ============================================================================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)